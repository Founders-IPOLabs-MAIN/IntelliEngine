from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body, Form
from motor.motor_asyncio import AsyncIOMotorGridFSBucket
from shared import (db, fs_bucket, logger, User, get_current_user, promote_new_user,
    datetime, timezone, uuid, io, os, ObjectId)
from valuation_engine import (
    calculate_dcf_valuation as vengine_dcf,
    calculate_nav_valuation, calculate_comparable_valuation,
    calculate_ddm_valuation, calculate_financial_ratios,
    compute_weighted_valuation
)

router = APIRouter()


@router.post("/valuation/projects")
async def create_valuation_project(data: dict = Body(...), user: User = Depends(get_current_user)):
    """Create a new valuation project"""
    valuation_id = f"val_{uuid.uuid4().hex[:12]}"
    doc = {
        "valuation_id": valuation_id,
        "user_id": user.user_id,
        "status": "draft",
        "current_step": 1,
        "company_profile": data.get("company_profile", {}),
        "documents": [],
        "financial_data": {},
        "valuation_config": {},
        "results": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.valuation_projects.insert_one(doc)
    await promote_new_user(user.user_id)
    return {"valuation_id": valuation_id, "status": "draft"}

@router.get("/valuation/projects")
async def list_valuation_projects(user: User = Depends(get_current_user)):
    """List user's valuation projects"""
    projects = await db.valuation_projects.find(
        {"user_id": user.user_id}, {"_id": 0}
    ).sort("updated_at", -1).to_list(50)
    return {"projects": projects}

@router.get("/valuation/projects/{valuation_id}")
async def get_valuation_project(valuation_id: str, user: User = Depends(get_current_user)):
    """Get a specific valuation project"""
    project = await db.valuation_projects.find_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Valuation project not found")
    return project

@router.put("/valuation/projects/{valuation_id}")
async def update_valuation_project(valuation_id: str, data: dict = Body(...), user: User = Depends(get_current_user)):
    """Save/update valuation project step data"""
    project = await db.valuation_projects.find_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Valuation project not found")

    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ["company_profile", "financial_data", "valuation_config", "current_step", "status"]:
        if field in data:
            update_fields[field] = data[field]

    await db.valuation_projects.update_one(
        {"valuation_id": valuation_id},
        {"$set": update_fields}
    )
    return {"status": "saved", "valuation_id": valuation_id}

@router.delete("/valuation/projects/{valuation_id}")
async def delete_valuation_project(valuation_id: str, user: User = Depends(get_current_user)):
    """Delete a valuation project"""
    result = await db.valuation_projects.delete_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"status": "deleted"}

@router.post("/valuation/projects/{valuation_id}/upload")
async def upload_valuation_document(
    valuation_id: str,
    file: UploadFile = File(...),
    doc_type: str = Form("financial_statement"),
    user: User = Depends(get_current_user)
):
    """Upload a document to a valuation project"""
    project = await db.valuation_projects.find_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Validate file type
    allowed = [".pdf", ".xlsx", ".xls", ".docx", ".doc", ".csv"]
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed:
        raise HTTPException(status_code=400, detail=f"File type {ext} not supported. Allowed: {', '.join(allowed)}")

    # Read and store in GridFS
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10MB limit
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")

    grid_fs = AsyncIOMotorGridFSBucket(db)
    grid_id = await grid_fs.upload_from_stream(file.filename, io.BytesIO(content))

    doc_record = {
        "filename": file.filename,
        "doc_type": doc_type,
        "gridfs_id": str(grid_id),
        "size": len(content),
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }

    await db.valuation_projects.update_one(
        {"valuation_id": valuation_id},
        {"$push": {"documents": doc_record}, "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"status": "uploaded", "document": doc_record}

@router.post("/valuation/projects/{valuation_id}/extract")
async def extract_financial_data(valuation_id: str, user: User = Depends(get_current_user)):
    """Use AI (GPT-5.2) to extract financial data from uploaded documents"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    import io

    project = await db.valuation_projects.find_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not project.get("documents"):
        raise HTTPException(status_code=400, detail="No documents uploaded yet")

    company = project.get("company_profile", {})
    currency = company.get("currency", "crores")

    # Try to read actual content from uploaded files
    file_contents = []
    grid_fs = AsyncIOMotorGridFSBucket(db)

    for doc in project["documents"][:3]:  # Max 3 docs
        try:
            from bson import ObjectId
            grid_id = ObjectId(doc["gridfs_id"])
            data_stream = io.BytesIO()
            await grid_fs.download_to_stream(grid_id, data_stream)
            data_stream.seek(0)
            fname = doc["filename"].lower()

            if fname.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(data_stream, data_only=True)
                sheets_text = []
                for sheet_name in wb.sheetnames[:5]:
                    ws = wb[sheet_name]
                    rows = []
                    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 60), values_only=True):
                        row_str = " | ".join([str(c) if c is not None else "" for c in row])
                        if row_str.strip(" |"):
                            rows.append(row_str)
                    if rows:
                        sheets_text.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
                file_contents.append(f"=== File: {doc['filename']} ===\n" + "\n\n".join(sheets_text))

            elif fname.endswith(".csv"):
                text = data_stream.read().decode("utf-8", errors="ignore")
                lines = text.strip().split("\n")[:60]
                file_contents.append(f"=== File: {doc['filename']} ===\n" + "\n".join(lines))

            elif fname.endswith(".pdf"):
                file_contents.append(f"=== File: {doc['filename']} (PDF - text extraction limited) ===\nPlease review manually.")

        except Exception as e:
            logger.error(f"Failed to read {doc['filename']}: {e}")

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    chat = LlmChat(
        api_key=api_key,
        session_id=f"val_extract_{uuid.uuid4().hex[:8]}",
        system_message="""You are a financial data extraction expert for Indian companies. 
Extract financial data from the provided spreadsheet/document content and return ONLY valid JSON.
All monetary values should be in the unit specified (Crores or Lakhs).
Match data to the closest fiscal year labels (FY2024, FY2023, etc.).
If a value is not found, use 0. Be precise with numbers."""
    ).with_model("openai", "gpt-5.2")

    doc_text = "\n\n".join(file_contents) if file_contents else "No readable file content available."

    prompt = f"""Extract financial data for {company.get('company_name', 'the company')} ({company.get('industry', 'general')} sector).
Currency unit: {currency}

UPLOADED DOCUMENT CONTENT:
{doc_text[:8000]}

Extract and return ONLY this JSON structure (no markdown, no extra text):
{{
  "years_data": [
    {{
      "year": "FY2024",
      "revenue": 0,
      "ebitda": 0,
      "pat": 0,
      "depreciation": 0,
      "capex": 0,
      "working_capital_change": 0,
      "total_assets": 0,
      "total_liabilities": 0,
      "net_worth": 0,
      "total_debt": 0,
      "cash_equivalents": 0,
      "current_assets": 0,
      "current_liabilities": 0,
      "dividend_per_share": 0
    }}
  ],
  "shares_outstanding": 100000,
  "face_value": 10,
  "notes": "Key observations"
}}

Include data for all years found in the documents. Map Revenue/Sales/Turnover to 'revenue', Profit After Tax/Net Profit to 'pat', Shareholders Equity to 'net_worth'."""

    try:
        response = await chat.send_message(UserMessage(text=prompt))
        import json
        json_str = response
        if "```json" in json_str:
            json_str = json_str.split("```json")[1].split("```")[0]
        elif "```" in json_str:
            json_str = json_str.split("```")[1].split("```")[0]
        extracted = json.loads(json_str.strip())
    except Exception as e:
        logger.error(f"AI extraction failed: {e}")
        extracted = {
            "years_data": [
                {"year": f"FY{2024-i}", "revenue": 0, "ebitda": 0, "pat": 0, "depreciation": 0,
                 "capex": 0, "working_capital_change": 0, "total_assets": 0, "total_liabilities": 0,
                 "net_worth": 0, "total_debt": 0, "cash_equivalents": 0, "current_assets": 0,
                 "current_liabilities": 0, "dividend_per_share": 0}
                for i in range(3)
            ],
            "shares_outstanding": 100000,
            "face_value": 10,
            "notes": "AI extraction did not succeed — please fill in data manually from your financial statements."
        }

    return {"extracted_data": extracted}

@router.post("/valuation/projects/{valuation_id}/calculate")
async def run_valuation_calculation(valuation_id: str, user: User = Depends(get_current_user)):
    """Run all selected valuation methods and generate comprehensive results"""
    from emergentintegrations.llm.chat import LlmChat, UserMessage

    project = await db.valuation_projects.find_one(
        {"valuation_id": valuation_id, "user_id": user.user_id}, {"_id": 0}
    )
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    financial_data = project.get("financial_data", {})
    config = project.get("valuation_config", {})
    company = project.get("company_profile", {})
    methods = config.get("methods", ["dcf"])

    if not financial_data.get("years_data"):
        raise HTTPException(status_code=400, detail="Financial data not provided")

    # Run valuation methods
    results = {}
    if "dcf" in methods:
        results["dcf"] = vengine_dcf(financial_data, config.get("dcf_config", {}))
    if "nav" in methods:
        results["nav"] = calculate_nav_valuation(financial_data, config.get("nav_config", {}))
    if "comparable" in methods:
        results["comparable"] = calculate_comparable_valuation(financial_data, config.get("comparable_config", {}))
    if "ddm" in methods:
        results["ddm"] = calculate_ddm_valuation(financial_data, config.get("ddm_config", {}))

    # Financial ratios
    ratios = calculate_financial_ratios(financial_data)

    # Default weights
    default_weights = {"dcf": 0.5, "nav": 0.25, "comparable": 0.25, "ddm": 0}
    weights = config.get("weights", default_weights)
    weighted = compute_weighted_valuation(results, weights)

    # Per-share value
    shares = financial_data.get("shares_outstanding", 1)
    per_share = round(weighted["weighted_average"] / shares, 2) if shares > 0 else 0

    # AI-powered analysis: Regulatory, Risk, Tax, Quality Control
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    ai_analysis = {}

    try:
        # Build context
        purpose = company.get("purpose", "ipo")
        purpose_labels = {"ma": "Merger & Acquisition", "esop": "ESOP Pricing", "ipo": "IPO",
                          "family_settlement": "Family Settlement", "tax_assessment": "Tax Assessment"}
        purpose_label = purpose_labels.get(purpose, purpose)

        latest = financial_data["years_data"][-1]
        val_summary = "\n".join([f"- {m.upper()}: ₹{r.get('valuation', 0):.2f} Cr" for m, r in results.items() if r.get("valuation", 0) > 0])

        # --- REGULATORY FRAMEWORK ---
        reg_chat = LlmChat(api_key=api_key, session_id=f"val_reg_{uuid.uuid4().hex[:6]}",
            system_message="You are an expert on Indian valuation regulations. Cite specific section/rule numbers."
        ).with_model("openai", "gpt-5.2")

        reg_prompt = f"""Analyze regulatory compliance for this valuation:
Purpose: {purpose_label}
Company: {company.get('company_name', 'N/A')} in {company.get('industry', 'N/A')} sector
Valuation Range: ₹{weighted['value_range_low']:.2f} - ₹{weighted['value_range_high']:.2f} Crores

Provide a concise analysis (max 200 words) covering:
1. Applicable Indian laws and specific section numbers
2. Whether registered valuer under Section 247 is mandatory
3. Key regulatory requirements for this transaction type
4. Any specific compliance flags

Use bullet points. Be precise with law citations."""

        ai_analysis["regulatory"] = await reg_chat.send_message(UserMessage(text=reg_prompt))

        # --- RISK ASSESSMENT ---
        risk_chat = LlmChat(api_key=api_key, session_id=f"val_risk_{uuid.uuid4().hex[:6]}",
            system_message="You are a business risk assessment expert for Indian companies."
        ).with_model("openai", "gpt-5.2")

        risk_prompt = f"""Assess valuation risks for:
Company: {company.get('company_name', 'N/A')} ({company.get('industry', 'N/A')})
Revenue: ₹{latest.get('revenue', 0):.2f} Cr, PAT: ₹{latest.get('pat', 0):.2f} Cr
Debt/Equity: {ratios.get('debt_equity', 'N/A')}, Current Ratio: {ratios.get('current_ratio', 'N/A')}
ROE: {ratios.get('roe', 'N/A')}%

Provide risk assessment in JSON format ONLY (no markdown):
{{"industry_risk": {{"level": "Low/Medium/High", "factors": ["factor1"]}},
"financial_risk": {{"level": "Low/Medium/High", "factors": ["factor1"]}},
"management_risk": {{"level": "Low/Medium/High", "factors": ["factor1"]}},
"legal_risk": {{"level": "Low/Medium/High", "factors": ["factor1"]}},
"market_risk": {{"level": "Low/Medium/High", "factors": ["factor1"]}},
"overall_rating": "Low/Medium/High",
"recommended_discount": 15,
"mitigation": ["suggestion1"]}}"""

        risk_raw = await risk_chat.send_message(UserMessage(text=risk_prompt))
        try:
            import json
            risk_str = risk_raw
            if "```json" in risk_str:
                risk_str = risk_str.split("```json")[1].split("```")[0]
            elif "```" in risk_str:
                risk_str = risk_str.split("```")[1].split("```")[0]
            ai_analysis["risk"] = json.loads(risk_str.strip())
        except Exception:
            ai_analysis["risk"] = {"overall_rating": "Medium", "recommended_discount": 15,
                                    "raw_analysis": risk_raw}

        # --- TAX IMPLICATIONS ---
        tax_chat = LlmChat(api_key=api_key, session_id=f"val_tax_{uuid.uuid4().hex[:6]}",
            system_message="You are an Indian tax expert specializing in corporate transactions and capital gains."
        ).with_model("openai", "gpt-5.2")

        tax_prompt = f"""Calculate tax implications for:
Transaction: {purpose_label}
Company Valuation: ₹{weighted['weighted_average']:.2f} Crores
Company Type: {company.get('company_type', 'private_limited')}

Provide concise analysis (max 200 words) covering:
1. Applicable capital gains tax (LTCG vs STCG)
2. Section 56(2)(viib) implications if applicable
3. TDS obligations
4. MAT/AMT considerations
5. Tax-efficient structuring suggestions

Use bullet points with section references."""

        ai_analysis["tax"] = await tax_chat.send_message(UserMessage(text=tax_prompt))

        # --- QUALITY CONTROL & CONFIDENCE ---
        qc_chat = LlmChat(api_key=api_key, session_id=f"val_qc_{uuid.uuid4().hex[:6]}",
            system_message="You are a valuation quality control expert. Provide confidence scores and validation."
        ).with_model("openai", "gpt-5.2")

        qc_prompt = f"""Review this valuation output:
Methods Used: {', '.join(methods)}
Results:
{val_summary}
Weighted Average: ₹{weighted['weighted_average']:.2f} Crores
Value Range: ₹{weighted['value_range_low']:.2f} - ₹{weighted['value_range_high']:.2f} Crores
Financial Ratios: ROE={ratios.get('roe', 'N/A')}%, D/E={ratios.get('debt_equity', 'N/A')}, Revenue CAGR={ratios.get('revenue_cagr', 'N/A')}%

Return ONLY JSON (no markdown):
{{"confidence_score": 75,
"variance_check": "PASS/FLAG",
"consistency_check": "PASS/FLAG",
"data_quality": "High/Medium/Low",
"flags": ["any concerns"],
"recommendations": ["suggestions"],
"executive_summary": "2-3 sentence summary of the valuation conclusion"}}"""

        qc_raw = await qc_chat.send_message(UserMessage(text=qc_prompt))
        try:
            import json
            qc_str = qc_raw
            if "```json" in qc_str:
                qc_str = qc_str.split("```json")[1].split("```")[0]
            elif "```" in qc_str:
                qc_str = qc_str.split("```")[1].split("```")[0]
            ai_analysis["quality_control"] = json.loads(qc_str.strip())
        except Exception:
            ai_analysis["quality_control"] = {"confidence_score": 70, "executive_summary": qc_raw}

    except Exception as e:
        logger.error(f"AI valuation analysis failed: {e}")
        ai_analysis["error"] = str(e)
        if "quality_control" not in ai_analysis:
            ai_analysis["quality_control"] = {"confidence_score": 60, "executive_summary": "AI analysis partially unavailable."}

    # Confidence score
    confidence = ai_analysis.get("quality_control", {}).get("confidence_score", 65)

    # Assemble final results
    final_results = {
        "methods": results,
        "ratios": ratios,
        "weighted_valuation": weighted,
        "per_share_value": per_share,
        "shares_outstanding": shares,
        "ai_analysis": ai_analysis,
        "confidence_score": confidence,
        "completed_at": datetime.now(timezone.utc).isoformat()
    }

    await db.valuation_projects.update_one(
        {"valuation_id": valuation_id},
        {"$set": {"results": final_results, "status": "completed",
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    return {"valuation_id": valuation_id, "results": final_results}
